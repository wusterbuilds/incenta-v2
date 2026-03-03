import openpyxl
import copy
import random
from datetime import datetime, timedelta

INPUT_FILE = "Denver 02172026 underwriting model - MZ.xlsx"
OUTPUT_FILE = "proforma_template_MASKED.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)

def is_formula(cell):
    return cell.data_type == 'f'

def is_number(val):
    return isinstance(val, (int, float)) and not isinstance(val, bool)

def round_nice(val):
    """Round to a 'clean' fake number appropriate to magnitude."""
    if val == 0:
        return 0
    av = abs(val)
    if av >= 1_000_000:
        return round(val / 500_000) * 500_000
    if av >= 100_000:
        return round(val / 50_000) * 50_000
    if av >= 10_000:
        return round(val / 5_000) * 5_000
    if av >= 1_000:
        return round(val / 500) * 500
    if av >= 100:
        return round(val / 50) * 50
    if av >= 10:
        return round(val / 5) * 5
    if av >= 1:
        return round(val)
    return round(val, 4)


# ---------------------------------------------------------------------------
# SPECIFIC OVERRIDES — key assumption cells replaced with plausible fakes
#   that keep the model structurally sound.  Everything below is intentionally
#   fictional and uses round numbers.
# ---------------------------------------------------------------------------

specific = {}

# ===== Unit Mix =====
um = "Unit Mix"
specific[(um, "B4")] = "1234 Example Blvd, Anytown, CO"
specific[(um, "B5")] = 200   # total units

# Scenario 1-4 unit counts (must sum to 200 for scenario 1)
for col in ["D", "F", "H", "J"]:
    specific[(um, f"{col}14")] = 140   # studio
    specific[(um, f"{col}15")] = 60    # 1b/1b
    specific[(um, f"{col}16")] = 0     # 2b/1b (some scenarios have 40)
for col in ["F"]:
    specific[(um, "F14")] = 100
    specific[(um, "F15")] = 60
    specific[(um, "F16")] = 40

# Unit sizes
for col in ["E", "G", "I", "K"]:
    specific[(um, f"{col}14")] = 350   # studio sf
    specific[(um, f"{col}15")] = 550   # 1b sf
    if (um, f"{col}16") not in specific:
        pass  # leave 0 / formula

specific[(um, "G14")] = 400
specific[(um, "G15")] = 550
specific[(um, "G16")] = 300

# Rents per unit type  (Scenarios 1-4, columns L-S)
for col_pair in [("L","M"), ("N","O"), ("P","Q"), ("R","S")]:
    specific[(um, f"{col_pair[0]}14")] = 1400   # studio rent
    specific[(um, f"{col_pair[0]}15")] = 1800   # 1b rent
    specific[(um, f"{col_pair[0]}16")] = 0
    specific[(um, f"{col_pair[0]}17")] = 0
    specific[(um, f"{col_pair[0]}18")] = 0

# ===== Lease UP =====
lu = "Lease UP"
specific[(lu, "F10")] = 200   # apartment units
specific[(lu, "G11")] = 200   # hotel rooms
specific[(lu, "H14")] = 75    # Hotel ADR

# Per-unit fee amounts
specific[(lu, "F39")] = 150   # admin fee
specific[(lu, "F40")] = 5     # credit builder
specific[(lu, "F41")] = 50    # application fee
specific[(lu, "F42")] = 50    # pet fee
specific[(lu, "F43")] = 10    # pet rent
specific[(lu, "F44")] = 5     # pest control
specific[(lu, "F45")] = 3     # laundry
specific[(lu, "F46")] = 40    # parking
specific[(lu, "F47")] = 10    # RLL liability
specific[(lu, "F48")] = 0     # W/D premium
specific[(lu, "F49")] = 10    # other income
specific[(lu, "F53")] = 0     # furniture rebill
specific[(lu, "F54")] = 0     # electric rebill
specific[(lu, "F55")] = 0     # internet
specific[(lu, "F56")] = 80    # utility bundle
specific[(lu, "F57")] = 0     # utility admin

# Expense per-unit / ratios (Lease Up)
specific[(lu, "F64")] = 2500  # tax per unit
specific[(lu, "F65")] = 800   # insurance per unit
specific[(lu, "F66")] = 0.06  # utility % of EGI
specific[(lu, "F67")] = 0.01
specific[(lu, "F68")] = 0.02
specific[(lu, "F69")] = 0.02
specific[(lu, "F70")] = 0.06
specific[(lu, "F71")] = 0.03
specific[(lu, "F73")] = 0.015
specific[(lu, "F74")] = 0.015
specific[(lu, "F27")] = 0.02  # bad debt
specific[(lu, "G65")] = 500
specific[(lu, "G66")] = 2000
specific[(lu, "G67")] = 800
specific[(lu, "G68")] = 5000
specific[(lu, "G69")] = 1000
specific[(lu, "G70")] = 3000
specific[(lu, "G71")] = 4000
specific[(lu, "G72")] = 200
specific[(lu, "G73")] = 5000
specific[(lu, "G74")] = 0

# Debt / construction line items
specific[(lu, "F83")] = 4000000
specific[(lu, "G83")] = 0.07
specific[(lu, "G85")] = 0.07
for i, mo_bal in enumerate([800000, 1600000, 2400000, 3200000, 4000000,
                             4800000, 5600000, 6400000, 7200000, 8000000], start=0):
    col_letter = chr(ord('H') + i)
    specific[(lu, f"{col_letter}84")] = mo_bal

specific[(lu, "F88")] = 0.02  # asset mgmt fee ratio
specific[(lu, "F94")] = 0.05  # reserve %

# Accounting & legal monthly values
for col in "HIJKLMNOPQRSTUVWXYZAAABACADAEAFAGAHAIAJAKALAMANAOAPAQARASATAUAVAW":
    if len(col) <= 2:
        specific[(lu, f"{col}89")] = 1200

# ===== Stable (annual) =====
st = "Stable"

# Purchase prices (4 scenarios: cols F, H, J, L)
specific[(st, "F14")] = 7000000
specific[(st, "H14")] = 5000000
specific[(st, "J14")] = 5000000
specific[(st, "L14")] = 5000000

# Impact fee
for c in ["F","H","J","L"]:
    specific[(st, f"{c}15")] = 0

# Renovation budget
specific[(st, "F16")] = 600000
specific[(st, "H16")] = 1800000
specific[(st, "J16")] = 2000000
specific[(st, "L16")] = 2000000

# LTV
for c in ["F","H","J","L"]:
    specific[(st, f"{c}20")] = 0.75

# Closing cost detail percentages
for c_pair in [("G",), ("I",), ("K",), ("M",)]:
    c = c_pair[0]
    specific[(st, f"{c}28")] = 0.005
    specific[(st, f"{c}29")] = 0.01
    specific[(st, f"{c}32")] = 0.03
    specific[(st, f"{c}33")] = 0.01

# Closing cost worksheet items
specific[(st, "Q16")] = 8000
specific[(st, "Q17")] = 35     # per-unit inspection
specific[(st, "Q18")] = 6000
specific[(st, "Q19")] = 1500
specific[(st, "Q20")] = 3500
specific[(st, "Q21")] = 0
specific[(st, "Q22")] = 40000
specific[(st, "Q24")] = 1200
specific[(st, "Q25")] = 150
specific[(st, "Q26")] = 400
specific[(st, "Q27")] = 300

# Property info
specific[(st, "V15")] = "1234 Example Blvd, Anytown, CO"

# Income assumptions - percentages
for c in ["G","I","K","M"]:
    specific[(st, f"{c}46")] = 0         # loss to lease
    specific[(st, f"{c}47")] = 0.05      # vacancy
    specific[(st, f"{c}48")] = 0.03      # collection loss
    specific[(st, f"{c}49")] = 0.01      # concessions
    specific[(st, f"{c}50")] = 0.005     # non-revenue units

# Expense line items (Scenario 1 = col F)
specific[(st, "F55")] = 175000    # taxes
specific[(st, "F56")] = 350000    # insurance
specific[(st, "F57")] = 400000    # utilities
specific[(st, "F58")] = 80000     # admin
specific[(st, "F59")] = 80000     # marketing
specific[(st, "F60")] = 80000     # turnover
specific[(st, "F61")] = 250000    # payroll
specific[(st, "F62")] = 150000    # PM
specific[(st, "F63")] = 75000     # contract services
specific[(st, "F64")] = 250000    # maint/repairs
specific[(st, "F65")] = 10000     # reserves

# Scenario 2 expenses (col H)
specific[(st, "H55")] = 175000
specific[(st, "H57")] = 130000
specific[(st, "H61")] = 130000

# Scenario 4 expenses (col L)
specific[(st, "L57")] = 100000
specific[(st, "L60")] = 20000
specific[(st, "L61")] = 110000
specific[(st, "L65")] = 20000
specific[(st, "L76")] = 4500000

# Reimbursable expenses
for c in ["F","H","J","L"]:
    specific[(st, f"{c}52")] = 0

# 1st Mortgage financing
specific[(st, "F77")] = 0.09      # interest rate
specific[(st, "H77")] = 0.07
specific[(st, "J77")] = 0.10
specific[(st, "L77")] = 0.10
for c in ["F","H","J","L"]:
    specific[(st, f"{c}78")] = 200     # amort years (IO indicator)
    specific[(st, f"{c}79")] = 12      # IO period months

# 2nd Mortgage
specific[(st, "F84")] = 0
specific[(st, "H84")] = 800000
specific[(st, "J84")] = 0
specific[(st, "L84")] = 800000
for c in ["F","H","J","L"]:
    specific[(st, f"{c}85")] = 0.06
    specific[(st, f"{c}86")] = 25
    specific[(st, f"{c}87")] = 6

# Refinance assumptions
specific[(st, "F92")] = 1
specific[(st, "H92")] = 2
specific[(st, "J92")] = 4
specific[(st, "L92")] = 2
for c in ["F","H","J","L"]:
    specific[(st, f"{c}94")] = 0.06
    specific[(st, f"{c}96")] = 0.055
    specific[(st, f"{c}97")] = 30
    specific[(st, f"{c}98")] = 0
    specific[(st, f"{c}99")] = 0.75

# Organic growth rates (Stable right-side panel)
specific[(st, "W44")] = 0
specific[(st, "X44")] = 0.03
specific[(st, "Y44")] = 0.03
specific[(st, "W45")] = 0
specific[(st, "X45")] = 0.03

# ===== Stable Monthly =====
sm = "Stable Monthly"

# Mirror the main financing and expense structure for the monthly model
specific[(sm, "F55")] = 125000
specific[(sm, "H55")] = 125000
specific[(sm, "J55")] = 125000
specific[(sm, "L55")] = 125000
specific[(sm, "P55")] = 500000

specific[(sm, "F57")] = 500000
specific[(sm, "H57")] = 500000
specific[(sm, "J57")] = 500000
specific[(sm, "L57")] = 500000
specific[(sm, "P57")] = 220000

specific[(sm, "F58")] = 30000
specific[(sm, "H58")] = 30000
specific[(sm, "J58")] = 30000
specific[(sm, "L58")] = 30000
specific[(sm, "P58")] = 30000

specific[(sm, "F59")] = 45000
specific[(sm, "H59")] = 45000
specific[(sm, "J59")] = 45000
specific[(sm, "L59")] = 45000
specific[(sm, "P59")] = 45000

specific[(sm, "F60")] = 60000
specific[(sm, "H60")] = 60000
specific[(sm, "J60")] = 60000
specific[(sm, "L60")] = 60000
specific[(sm, "P60")] = 60000

specific[(sm, "F61")] = 200000
specific[(sm, "H61")] = 200000
specific[(sm, "J61")] = 200000
specific[(sm, "L61")] = 200000
specific[(sm, "P61")] = 220000

specific[(sm, "F62")] = 90000
specific[(sm, "H62")] = 90000
specific[(sm, "J62")] = 90000
specific[(sm, "L62")] = 90000
specific[(sm, "P62")] = 90000

specific[(sm, "F63")] = 75000
specific[(sm, "H63")] = 75000
specific[(sm, "J63")] = 75000
specific[(sm, "L63")] = 75000
specific[(sm, "P63")] = 0

specific[(sm, "F64")] = 35000
specific[(sm, "H64")] = 35000
specific[(sm, "J64")] = 35000
specific[(sm, "L64")] = 35000
specific[(sm, "P64")] = 50000

specific[(sm, "F65")] = 120000
specific[(sm, "H65")] = 120000
specific[(sm, "J65")] = 120000
specific[(sm, "L65")] = 120000
specific[(sm, "P65")] = 45000

# Vacancy/loss percentages in Stable Monthly
for c in ["G","I","K","M"]:
    specific[(sm, f"{c}46")] = 0
    specific[(sm, f"{c}47")] = 0.04
    specific[(sm, f"{c}48")] = 0.02
    specific[(sm, f"{c}49")] = 0
    specific[(sm, f"{c}50")] = 0
    specific[(sm, f"{c}32")] = 0.03
    specific[(sm, f"{c}33")] = 0.02

# 1st Mortgage (monthly model)
specific[(sm, "F77")] = 0.06
specific[(sm, "H77")] = 0.07
specific[(sm, "J77")] = 0.06
specific[(sm, "L77")] = 0.06
for c in ["F","H","J","L"]:
    specific[(sm, f"{c}78")] = 200
    specific[(sm, f"{c}79")] = 24

# 2nd Mortgage (monthly)
specific[(sm, "F84")] = 0
specific[(sm, "H84")] = 2500000
specific[(sm, "J84")] = 0
specific[(sm, "L84")] = 0
for c in ["F","H","J","L"]:
    specific[(sm, f"{c}85")] = 0.06
    specific[(sm, f"{c}86")] = 30
    specific[(sm, f"{c}87")] = 6

# Refinance (monthly)
for c in ["F","H","J","L"]:
    specific[(sm, f"{c}92")] = 30     # month
    specific[(sm, f"{c}94")] = 0.065
    specific[(sm, f"{c}96")] = 0.06
    specific[(sm, f"{c}97")] = 30
    specific[(sm, f"{c}98")] = 60
    specific[(sm, f"{c}99")] = 0.8

# Organic growth rates (monthly panel)
specific[(sm, "W44")] = 0
specific[(sm, "X44")] = 0
specific[(sm, "W45")] = 0
specific[(sm, "X45")] = 0

# JV/seller assumptions
specific[(sm, "Q35")] = 0
specific[(sm, "Q36")] = 0
specific[(sm, "Q37")] = 0

# ===== CoC sheets — mostly formulas, but override waterfall split inputs =====
coc = "CoC from Operations"
cocm = "CoC from Operations Monthly"
# These are almost entirely formulas; the input cells for preferred return
# and split ratios are pulled from Stable. No additional overrides needed.

# ---------------------------------------------------------------------------
# APPLY OVERRIDES AND GENERIC MASKING
# ---------------------------------------------------------------------------

masked_count = 0
formula_preserved = 0
label_preserved = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue

            key = (sheet_name, cell.coordinate)

            # If we have a specific override, apply it
            if key in specific:
                cell.value = specific[key]
                masked_count += 1
                continue

            # Skip formulas — preserve them exactly
            if is_formula(cell):
                formula_preserved += 1
                continue

            # Skip strings/labels
            if isinstance(cell.value, str):
                label_preserved += 1
                continue

            # Skip None
            if cell.value is None:
                continue

            # Skip datetime values (just shift by a random offset)
            if isinstance(cell.value, datetime):
                cell.value = cell.value.replace(year=2025, month=6, day=15)
                masked_count += 1
                continue

            # For remaining hard-coded numbers, apply generic rounding/masking
            if is_number(cell.value):
                val = cell.value
                # Preserve 0
                if val == 0:
                    continue
                # Percentages (0 < |x| < 1) — keep as-is, these are rates
                if 0 < abs(val) < 1:
                    # Round to nearest 0.005 to mask slightly
                    cell.value = round(round(val * 200) / 200, 4)
                    masked_count += 1
                    continue
                # Small integers that look like counts, months, years
                if abs(val) <= 500 and val == int(val):
                    cell.value = round_nice(val)
                    masked_count += 1
                    continue
                # Dollar amounts — apply scaling + rounding
                factor = random.uniform(0.75, 1.25)
                cell.value = round_nice(val * factor)
                masked_count += 1

wb.save(OUTPUT_FILE)

print(f"Masked file saved as: {OUTPUT_FILE}")
print(f"  Cells with specific overrides or generic masking: {masked_count}")
print(f"  Formulas preserved: {formula_preserved}")
print(f"  Labels preserved: {label_preserved}")
