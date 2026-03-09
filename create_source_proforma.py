"""
Generate a single-sheet "external" pro forma Excel file with a completely
different layout from proforma_template_MASKED.xlsx, but populated with the
same mock numbers.  This is the file the user uploads during the demo.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "sample_external_proforma.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pro Forma Summary"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="2F5233")
LABEL_FONT = Font(name="Calibri", size=10, color="333333")
VALUE_FONT = Font(name="Calibri", size=10, color="000000")
SUBLABEL_FONT = Font(name="Calibri", size=10, color="666666", italic=True)

HEADER_FILL = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
SECTION_FILL = PatternFill(start_color="E8F0E9", end_color="E8F0E9", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F7F9F7", end_color="F7F9F7", fill_type="solid")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)
SECTION_BORDER = Border(
    bottom=Side(style="medium", color="2F5233"),
)

DOLLAR_FMT = '#,##0'
DOLLAR_FMT_DEC = '#,##0.00'
PCT_FMT = '0.0%'
INT_FMT = '#,##0'

# Column widths
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 18

def write_header(row, text):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 32
    for c in range(3, 6):
        ws.cell(row=row, column=c).fill = HEADER_FILL

def write_section(row, text):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    for c in range(3, 6):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    return row

def write_row(row, label, value, fmt=None, sublabel=None, alt=False):
    lbl = ws.cell(row=row, column=2, value=label)
    lbl.font = SUBLABEL_FONT if sublabel else LABEL_FONT
    lbl.border = THIN_BORDER
    val = ws.cell(row=row, column=3, value=value)
    val.font = VALUE_FONT
    val.alignment = Alignment(horizontal="right")
    val.border = THIN_BORDER
    if fmt:
        val.number_format = fmt
    if alt:
        for c in range(2, 6):
            ws.cell(row=row, column=c).fill = ALT_ROW_FILL

def write_table_header(row, cols):
    for i, col_name in enumerate(cols):
        cell = ws.cell(row=row, column=2 + i, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="2F5233")
        cell.border = SECTION_BORDER
        cell.alignment = Alignment(horizontal="right" if i > 0 else "left")

def write_table_row(row, values, fmts=None, alt=False):
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=2 + i, value=val)
        cell.font = VALUE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right" if i > 0 else "left")
        if fmts and i < len(fmts) and fmts[i]:
            cell.number_format = fmts[i]
        if alt:
            cell.fill = ALT_ROW_FILL

# ---------------------------------------------------------------------------
# Build the sheet
# ---------------------------------------------------------------------------
r = 2

write_header(r, "  INVESTMENT SUMMARY  —  1234 Quebec St, Denver, CO")
r += 2

# --- Project Overview ---
write_section(r, "PROJECT OVERVIEW")
r += 1
write_row(r, "Property Address", "1234 Quebec St, Denver, CO"); r += 1
write_row(r, "Property Type", "Multifamily — Hotel Conversion", alt=True); r += 1
write_row(r, "Total Units", 200, fmt=INT_FMT); r += 1
write_row(r, "Year Built / Converted", "1968 / 2025", alt=True); r += 1
write_row(r, "Market", "Denver Metro, CO"); r += 1

r += 1

# --- Acquisition ---
write_section(r, "ACQUISITION")
r += 1
write_row(r, "Purchase Price", 7_000_000, fmt=DOLLAR_FMT); r += 1
write_row(r, "Renovation Budget", 600_000, fmt=DOLLAR_FMT, alt=True); r += 1
write_row(r, "Total Project Cost", 7_600_000, fmt=DOLLAR_FMT); r += 1
write_row(r, "Price Per Unit", 35_000, fmt=DOLLAR_FMT, alt=True); r += 1
write_row(r, "Loan-to-Value", 0.75, fmt=PCT_FMT); r += 1

r += 1

# --- Unit Mix (table) ---
write_section(r, "UNIT MIX")
r += 1
write_table_header(r, ["Type", "Units", "Avg SF", "Rent/Mo", "Annual Revenue"])
r += 1
write_table_row(r, ["Studio", 140, 350, 1400, 140 * 1400 * 12],
                fmts=[None, INT_FMT, INT_FMT, DOLLAR_FMT, DOLLAR_FMT]); r += 1
write_table_row(r, ["1 Bed / 1 Bath", 60, 550, 1800, 60 * 1800 * 12],
                fmts=[None, INT_FMT, INT_FMT, DOLLAR_FMT, DOLLAR_FMT], alt=True); r += 1
write_table_row(r, ["Total / Wtd Avg", 200, 410, 1518, 200 * 1518 * 12],
                fmts=[None, INT_FMT, INT_FMT, DOLLAR_FMT, DOLLAR_FMT]); r += 1
# Bold the totals row
for c in range(2, 7):
    cell = ws.cell(row=r - 1, column=c)
    cell.font = Font(name="Calibri", size=10, bold=True, color="2F5233")

r += 1

# --- Operating Expenses ---
write_section(r, "OPERATING EXPENSES (Annual)")
r += 1
expenses = [
    ("Real Estate Taxes", 125_000),
    ("Utilities", 500_000),
    ("Administrative", 30_000),
    ("Marketing & Leasing", 45_000),
    ("Turnover / Make-Ready", 60_000),
    ("Payroll & Benefits", 200_000),
    ("Property Management", 90_000),
    ("Contract Services", 75_000),
    ("Maintenance & Repairs", 35_000),
    ("Replacement Reserves", 120_000),
]
total_exp = 0
for i, (label, amount) in enumerate(expenses):
    write_row(r, label, amount, fmt=DOLLAR_FMT, alt=(i % 2 == 1))
    total_exp += amount
    r += 1
write_row(r, "Total Operating Expenses", total_exp, fmt=DOLLAR_FMT)
ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, bold=True, color="2F5233")
ws.cell(row=r, column=3).font = Font(name="Calibri", size=10, bold=True, color="2F5233")
r += 1
write_row(r, "Expense Per Unit", total_exp / 200, fmt=DOLLAR_FMT, alt=True)
r += 2

# --- Financing ---
write_section(r, "FINANCING — 1st Mortgage")
r += 1
write_row(r, "Loan Amount (LTV 75%)", 5_250_000, fmt=DOLLAR_FMT); r += 1
write_row(r, "Interest Rate", 0.06, fmt=PCT_FMT, alt=True); r += 1
write_row(r, "Amortization", "Interest Only", alt=False); r += 1
write_row(r, "IO Period", "24 months", alt=True); r += 1
write_row(r, "2nd Mortgage", "None"); r += 1

r += 1

write_section(r, "REFINANCE ASSUMPTIONS")
r += 1
write_row(r, "Refinance Month", 30, fmt=INT_FMT); r += 1
write_row(r, "Stabilized Rate", 0.065, fmt=PCT_FMT, alt=True); r += 1
write_row(r, "Exit Rate", 0.06, fmt=PCT_FMT); r += 1
write_row(r, "Amortization (years)", 30, fmt=INT_FMT, alt=True); r += 1
write_row(r, "Refinance LTV", 0.80, fmt=PCT_FMT); r += 1

r += 2

# --- Disclaimer ---
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
cell = ws.cell(row=r, column=2, value="This pro forma was prepared for illustrative purposes. All projections are estimates.")
cell.font = Font(name="Calibri", size=9, color="999999", italic=True)

# Print setup
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

wb.save(OUTPUT_FILE)
print(f"Created: {OUTPUT_FILE}")
print(f"  Rows used: {r}")
