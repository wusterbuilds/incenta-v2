"""
Take proforma_template_MASKED.xlsx and zero-out / blank the key input cells
that populateTemplate() writes to, so the demo starts with a visibly empty
template.  All formulas, formatting, labels, and structure are preserved.
"""

import openpyxl

INPUT_FILE = "proforma_template_MASKED.xlsx"
OUTPUT_FILE = "proforma_template_EMPTY.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)

def is_formula(cell):
    return cell.data_type == 'f'

# ---------------------------------------------------------------------------
# Cells to clear — these match the CellChange[] in translate.ts mock data.
# Only hard-coded input cells are cleared; formulas are never touched.
# ---------------------------------------------------------------------------

cells_to_zero = {
    # Stable Monthly — Acquisition
    ("Stable Monthly", "V15"): "",            # address → blank string
    ("Stable Monthly", "F14"): 0,             # purchase price
    ("Stable Monthly", "F16"): 0,             # renovation budget
    ("Stable Monthly", "F20"): 0,             # LTV

    # Stable Monthly — Expenses
    ("Stable Monthly", "F55"): 0,             # taxes
    ("Stable Monthly", "F57"): 0,             # utilities
    ("Stable Monthly", "F58"): 0,             # admin
    ("Stable Monthly", "F59"): 0,             # marketing
    ("Stable Monthly", "F60"): 0,             # turnover
    ("Stable Monthly", "F61"): 0,             # payroll
    ("Stable Monthly", "F62"): 0,             # PM
    ("Stable Monthly", "F63"): 0,             # contract services
    ("Stable Monthly", "F64"): 0,             # maintenance
    ("Stable Monthly", "F65"): 0,             # reserves

    # Stable Monthly — Financing
    ("Stable Monthly", "F77"): 0,             # interest rate
    ("Stable Monthly", "F78"): 0,             # amortization
    ("Stable Monthly", "F79"): 0,             # IO months
    ("Stable Monthly", "F84"): 0,             # 2nd mortgage

    # Stable Monthly — Refinance
    ("Stable Monthly", "F92"): 0,             # refi month
    ("Stable Monthly", "F94"): 0,             # refi rate
    ("Stable Monthly", "F96"): 0,             # refi exit rate
    ("Stable Monthly", "F97"): 0,             # refi amort
    ("Stable Monthly", "F99"): 0,             # refi LTV

    # Unit Mix
    ("Unit Mix", "B5"):  0,                   # total units
    ("Unit Mix", "D14"): 0,                   # studio units
    ("Unit Mix", "D15"): 0,                   # 1BR units
    ("Unit Mix", "E14"): 0,                   # studio SF
    ("Unit Mix", "E15"): 0,                   # 1BR SF
    ("Unit Mix", "L14"): 0,                   # studio rent
    ("Unit Mix", "L15"): 0,                   # 1BR rent
}

# Also zero out the Stable (annual) sheet equivalents so the whole workbook
# looks consistently empty.  Skip any cell that is a formula.

stable_annual_zeros = {
    ("Stable", "V15"): "",
    ("Stable", "F14"): 0,
    ("Stable", "F16"): 0,
    ("Stable", "F20"): 0,
    ("Stable", "F55"): 0,
    ("Stable", "F56"): 0,
    ("Stable", "F57"): 0,
    ("Stable", "F58"): 0,
    ("Stable", "F59"): 0,
    ("Stable", "F60"): 0,
    ("Stable", "F61"): 0,
    ("Stable", "F62"): 0,
    ("Stable", "F63"): 0,
    ("Stable", "F64"): 0,
    ("Stable", "F65"): 0,
    ("Stable", "F77"): 0,
    ("Stable", "F78"): 0,
    ("Stable", "F79"): 0,
    ("Stable", "F84"): 0,
}

# Unit Mix columns for scenarios 1 (D) — also zero col F (scenario 2) units
unit_mix_extra = {
    ("Unit Mix", "B4"): "",               # address on unit mix
    ("Unit Mix", "D16"): 0,
    ("Unit Mix", "D17"): 0,
    ("Unit Mix", "D18"): 0,
    ("Unit Mix", "E16"): 0,
    ("Unit Mix", "E17"): 0,
    ("Unit Mix", "E18"): 0,
    ("Unit Mix", "L16"): 0,
    ("Unit Mix", "L17"): 0,
    ("Unit Mix", "L18"): 0,
    ("Unit Mix", "F14"): 0,
    ("Unit Mix", "F15"): 0,
    ("Unit Mix", "F16"): 0,
}

all_cells = {}
all_cells.update(cells_to_zero)
all_cells.update(stable_annual_zeros)
all_cells.update(unit_mix_extra)

cleared = 0
skipped_formula = 0

for (sheet_name, cell_ref), new_val in all_cells.items():
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: sheet '{sheet_name}' not found, skipping {cell_ref}")
        continue
    ws = wb[sheet_name]
    cell = ws[cell_ref]

    if is_formula(cell):
        skipped_formula += 1
        continue

    cell.value = new_val
    cleared += 1

wb.save(OUTPUT_FILE)
print(f"Created: {OUTPUT_FILE}")
print(f"  Cells cleared: {cleared}")
print(f"  Formulas preserved (skipped): {skipped_formula}")
